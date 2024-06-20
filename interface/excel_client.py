import re
from openpyxl import load_workbook
from consts import COMPLETE, NA, IN_PROGRESS


class ExcelClient:
    def __init__(self, wb_path):
        self.wb = load_workbook(wb_path, keep_vba=True)
        self.ws = self.wb['PSS Main']
        self.id_col = self.get_col_num_from_heading('JOB #')
        self.name_col = self.get_col_num_from_heading('PROJECT')
        self.sca_status_col = self.get_col_num_from_heading('STATUS SC/COR/AF')
        self.sca_status_date_col = self.get_col_num_from_heading('STATUS CHANGED')
        self.ground_status_col = self.get_col_num_from_heading('STATUS GND')
        self.ground_status_date_col = self.get_col_num_from_heading('STATUS CHANGED2')
        self.project_status_col = self.get_col_num_from_heading('PROJECT STATUS')
        self.project_data = []
        self.wb_path = wb_path

    def get_projects_data(self, start_row=0, end_row=0):
        projects_data = []
        if not end_row:
            end_row = self.ws.max_row
        for row in self.ws.iter_rows(min_row=start_row, max_row=end_row):
            id_cell = row[self.id_col]
            name_cell = row[self.name_col]
            sca_status_cell = row[self.sca_status_col]
            ground_status_cell = row[self.ground_status_col]
            project_status_cell = row[self.project_status_col]
            if id_cell.value and (project_status_cell.value == IN_PROGRESS
                                  and (sca_status_cell.value not in [COMPLETE, NA]
                                       or ground_status_cell.value not in [COMPLETE, NA])):
                search = re.search(r'^\d+', str(id_cell.value))
                projects_data.append({'row': id_cell.row,
                                      'id': search.group(),
                                      'name': name_cell.value,
                                      'sca_status': sca_status_cell.value,
                                      'sca_status_date': None,
                                      'ground_status': ground_status_cell.value,
                                      'ground_status_date': None})
        return projects_data

    def set_project_statuses(self, projects_data):
        for project_data in projects_data:
            row = self.ws[project_data['row']]
            row[self.sca_status_col].value = project_data['sca_status']
            row[self.ground_status_col].value = project_data['ground_status']
            row[self.sca_status_date_col].value = project_data['sca_status_date']
            row[self.ground_status_date_col].value = project_data['ground_status_date']
        self.wb.save(self.wb_path)
        self.wb.close()

    def get_col_num_from_heading(self, heading):
        for col in self.ws.iter_cols(1, self.ws.max_column):
            if col[5].value == heading:
                return col[5].column - 1
