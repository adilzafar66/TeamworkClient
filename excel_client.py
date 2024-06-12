import re
from openpyxl import load_workbook


class ExcelClient:
    def __init__(self, wb_path):
        self.wb = load_workbook(wb_path, keep_vba=True)
        self.ws = self.wb['PSS Main']
        self.id_col = 0
        self.sca_status_col = 15
        self.ground_status_col = 23
        self.project_status_col = 30
        self.project_data = []

    def get_projects_data(self, start_row=0, end_row=0):
        projects_data = []
        if not end_row:
            end_row = self.ws.max_row
        for row in self.ws.iter_rows(min_row=start_row, max_row=end_row):
            id_cell = row[self.id_col]
            sca_status_cell = row[self.sca_status_col]
            ground_status_cell = row[self.ground_status_col]
            project_status_cell = row[self.project_status_col]
            if id_cell.value and ('IN PROGRESS' in project_status_cell.value or
                                  'TO BE STARTED' in project_status_cell.value):
                search = re.search(r'^\d+', str(id_cell.value))
                projects_data.append({'row': id_cell.row,
                                      'id': search.group(),
                                      'sca_status': sca_status_cell.value,
                                      'ground_status': ground_status_cell.value})
        return projects_data

    def set_project_statuses(self, projects_data):
        for project_data in projects_data:
            row = self.ws[project_data['row']]
            row[self.id_col].value = project_data['id']
            row[self.sca_status_col] = project_data['sca_status']
            row[self.ground_status_col] = project_data['ground_status']
